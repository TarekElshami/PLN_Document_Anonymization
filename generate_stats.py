import os
from collections import defaultdict
from lxml import etree
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
import numpy as np


class MeddocanAnalyzer:
    """Clase para analizar y comparar anotaciones MEDDOCAN"""

    def __init__(self):
        self.tag_errors = defaultdict(lambda: {'fp': [], 'fn': [], 'tp': []})
        self.tag_sentences = defaultdict(int)
        self.results = {}
        self.confusion_matrices = {}

    def parse_i2b2_annotations(self, xml_file):
        """Parsea las anotaciones de un archivo XML i2b2"""
        annotations = []
        tree = etree.parse(xml_file)
        root = tree.getroot()

        text = root.find("TEXT").text if root.find("TEXT") is not None else ""

        for tag in root.find("TAGS"):
            annotation = {
                'type': tag.tag,
                'start': int(tag.get('start')),
                'end': int(tag.get('end')),
                'text': tag.get('text'),
                'tag_text': text[int(tag.get('start')):int(tag.get('end'))]
            }
            if 'TYPE' in tag.attrib:
                annotation['subtype'] = tag.get('TYPE')
            annotations.append(annotation)

        return {
            'text': text,
            'annotations': annotations
        }

    def get_context(self, text, start, end, window=30):
        """Obtiene el contexto alrededor de una anotación"""
        context_start = max(0, start - window)
        context_end = min(len(text), end + window)
        before = text[context_start:start]
        annotated = text[start:end]
        after = text[end:context_end]
        return f"...{before}>>>{annotated}<<<{after}..."

    def find_overlapping_annotation(self, gold_ann, system_annotations, tolerance=10):
        """
        Encuentra si hay alguna anotación del sistema que se solape con la anotación gold.
        Retorna la anotación del sistema que más se solape, o None si no hay solapamiento.

        Args:
            gold_ann: Anotación gold
            system_annotations: Lista de anotaciones del sistema
            tolerance: Tolerancia en caracteres para considerar solapamiento
        """
        best_match = None
        best_overlap = 0

        gold_start, gold_end = gold_ann['start'], gold_ann['end']

        for sys_ann in system_annotations:
            sys_start, sys_end = sys_ann['start'], sys_ann['end']

            # Calcular solapamiento
            overlap_start = max(gold_start, sys_start)
            overlap_end = min(gold_end, sys_end)

            if overlap_start < overlap_end:  # Hay solapamiento
                overlap_size = overlap_end - overlap_start
                if overlap_size > best_overlap:
                    best_overlap = overlap_size
                    best_match = sys_ann

            # También considerar proximidad (dentro de la tolerancia)
            elif (abs(gold_start - sys_start) <= tolerance or
                  abs(gold_end - sys_end) <= tolerance or
                  abs(gold_start - sys_end) <= tolerance or
                  abs(gold_end - sys_start) <= tolerance):
                if best_match is None:  # Solo si no hemos encontrado un solapamiento real
                    best_match = sys_ann

        return best_match

    def build_confusion_matrix_data(self, gold, system, filename):
        """
        Construye los datos para la matriz de confusión comparando anotaciones gold vs system
        """
        gold_ann = gold['annotations']
        sys_ann = system['annotations']
        confusion_data = []

        # Para cada anotación gold, determinar qué predijo el sistema
        for g_ann in gold_ann:
            gold_tag = g_ann.get('subtype', g_ann['type'])

            # Buscar coincidencia exacta primero (TP)
            exact_match = None
            for s_ann in sys_ann:
                if (s_ann['start'] == g_ann['start'] and
                        s_ann['end'] == g_ann['end'] and
                        s_ann.get('subtype', s_ann['type']) == gold_tag):
                    exact_match = s_ann
                    break

            if exact_match:
                # True Positive
                predicted_tag = exact_match.get('subtype', exact_match['type'])
                confusion_data.append({
                    'gold': gold_tag,
                    'predicted': predicted_tag,
                    'type': 'TP',
                    'document': filename,
                    'text': g_ann['text'],
                    'position': (g_ann['start'], g_ann['end'])
                })
            else:
                # Buscar solapamiento o proximidad
                overlapping_ann = self.find_overlapping_annotation(g_ann, sys_ann, 0)

                if overlapping_ann:
                    # Clasificación incorrecta (confusión entre tags)
                    predicted_tag = overlapping_ann.get('subtype', overlapping_ann['type'])
                    confusion_data.append({
                        'gold': gold_tag,
                        'predicted': predicted_tag,
                        'type': 'CONFUSION',
                        'document': filename,
                        'text': g_ann['text'],
                        'position': (g_ann['start'], g_ann['end']),
                        'system_text': overlapping_ann['text'],
                        'system_position': (overlapping_ann['start'], overlapping_ann['end'])
                    })

        return confusion_data

    def compare_document_annotations(self, gold, system, tag_errors, filename, out_f=None):
        """Compara las anotaciones de un documento (funcionalidad original mantenida)"""
        gold_ann = gold['annotations']
        sys_ann = system['annotations']
        false_positives = []
        false_negatives = []

        # Buscar True Positives y False Positives
        for s_ann in sys_ann:
            matched = False
            for g_ann in gold_ann:
                if (s_ann['start'] == g_ann['start'] and
                        s_ann['end'] == g_ann['end'] and
                        s_ann.get('subtype', None) == g_ann.get('subtype', None)):
                    matched = True
                    tag_type = g_ann.get('subtype', g_ann['type'])
                    tag_errors[tag_type]['tp'].append({
                        'text': g_ann['text'],
                        'position': (g_ann['start'], g_ann['end']),
                        'document': filename
                    })
                    break

            if not matched:
                tag_type = s_ann.get('subtype', s_ann['type'])
                tag_errors[tag_type]['fp'].append({
                    'text': system['text'][s_ann['start']:s_ann['end']],
                    'system_text': s_ann['text'],
                    'position': (s_ann['start'], s_ann['end']),
                    'document': filename
                })
                false_positives.append(s_ann)

        # Buscar False Negatives
        for g_ann in gold_ann:
            matched = False
            for s_ann in sys_ann:
                if (s_ann['start'] == g_ann['start'] and
                        s_ann['end'] == g_ann['end'] and
                        s_ann.get('subtype', None) == g_ann.get('subtype', None)):
                    matched = True
                    break
            if not matched:
                tag_type = g_ann.get('subtype', g_ann['type'])
                tag_errors[tag_type]['fn'].append({
                    'text': gold['text'][g_ann['start']:g_ann['end']],
                    'gold_text': g_ann['text'],
                    'position': (g_ann['start'], g_ann['end']),
                    'document': filename
                })
                false_negatives.append(g_ann)

        # Escribir errores si se proporciona archivo de salida
        if out_f:
            if false_positives or false_negatives:
                out_f.write(f"\nErrors found in document: {filename}\n")
                if false_positives:
                    out_f.write("\nFalse Positives (System tagged but shouldn't):\n")
                    for fp in false_positives:
                        out_f.write(f"- {fp.get('subtype', fp['type'])}: '{fp['text']}' "
                                    f"(pos: {fp['start']}-{fp['end']})\n")
                        out_f.write(f"  Context: '{self.get_context(system['text'], fp['start'], fp['end'])}'\n")

                if false_negatives:
                    out_f.write("\nFalse Negatives (System missed):\n")
                    for fn in false_negatives:
                        out_f.write(f"- {fn.get('subtype', fn['type'])}: '{fn['text']}' "
                                    f"(pos: {fn['start']}-{fn['end']})\n")
                        out_f.write(f"  Context: '{self.get_context(gold['text'], fn['start'], fn['end'])}'\n")
            else:
                out_f.write(f"No errors found in document: {filename}.\n")

        return false_positives, false_negatives

    def leak_score(self, fn, num_sentences):
        """Calcula el leak score"""
        try:
            return round(len(fn) / num_sentences, 4) if num_sentences > 0 else 0.0
        except Exception:
            return "NA"

    def compute_metrics(self, tag_errors, tag_sentences):
        """Computa las métricas por tag"""
        metrics = {}

        for tag, values in tag_errors.items():
            tp = len(values['tp'])
            fp = len(values['fp'])
            fn = len(values['fn'])

            precision = tp / (tp + fp) if (tp + fp) else 0.0
            recall = tp / (tp + fn) if (tp + fn) else 0.0
            f1 = (2 * precision * recall) / (precision + recall) if (precision + recall) else 0.0
            leak = self.leak_score(values['fn'], tag_sentences.get(tag, 0))

            metrics[tag] = {
                'precision': round(precision, 4),
                'recall': round(recall, 4),
                'f1': round(f1, 4),
                'leak': leak,
                'tp': tp,
                'fp': fp,
                'fn': fn
            }

        return metrics

    def create_confusion_matrix(self, confusion_data, prompt_name):
        """
        Crea la matriz de confusión a partir de los datos de confusión
        """
        # Obtener todos los tags únicos
        all_tags = set()
        for item in confusion_data:
            all_tags.add(item['gold'])
            all_tags.add(item['predicted'])

        final_tags = sorted(set(
            item['gold'] for item in confusion_data
        ).union(
            item['predicted'] for item in confusion_data
        ))

        # Crear matriz
        matrix_size = len(final_tags)
        confusion_matrix = np.zeros((matrix_size, matrix_size), dtype=int)

        # Llenar la matriz
        for item in confusion_data:
            gold_idx = final_tags.index(item['gold'])
            pred_idx = final_tags.index(item['predicted'])
            confusion_matrix[gold_idx][pred_idx] += 1

        # Guardar en resultados
        self.confusion_matrices[prompt_name] = {
            'matrix': confusion_matrix,
            'labels': final_tags,
            'data': confusion_data
        }

        return confusion_matrix, final_tags

    def plot_confusion_matrix(self, prompt_name, output_dir, save_plot=True):
        """
        Crea y opcionalmente guarda el plot de la matriz de confusión
        """
        if prompt_name not in self.confusion_matrices:
            print(f"No confusion matrix data found for {prompt_name}")
            return None

        matrix_data = self.confusion_matrices[prompt_name]
        matrix = matrix_data['matrix']
        labels = matrix_data['labels']

        plt.figure(figsize=(12, 10))

        # Crear heatmap
        log_matrix = np.log1p(matrix)

        sns.heatmap(log_matrix,
                    annot=matrix,  # Mostrar los valores reales, no los logaritmos
                    fmt='d',
                    cmap='Blues',
                    xticklabels=labels,
                    yticklabels=labels,
                    cbar_kws={'label': 'Número de casos (escala logarítmica)'})

        plt.title(f'Matriz de Confusión - {prompt_name}', fontsize=16, pad=20)
        plt.xlabel('Predicción del Sistema', fontsize=12)
        plt.ylabel('Etiqueta Gold (Verdadera)', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.yticks(rotation=0)

        # Ajustar layout
        plt.tight_layout()

        if save_plot:
            output_path = Path(output_dir) / f'confusion_matrix_{prompt_name.lower().replace(" ", "_")}.png'
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"Matriz de confusión guardada en: {output_path}")

        plt.show()

        return matrix, labels

    def save_confusion_matrix_to_excel(self, prompt_name, output_dir):
        """
        Guarda la matriz de confusión y datos detallados en Excel
        """
        if prompt_name not in self.confusion_matrices:
            print(f"No confusion matrix data found for {prompt_name}")
            return None

        matrix_data = self.confusion_matrices[prompt_name]
        matrix = matrix_data['matrix']
        labels = matrix_data['labels']
        confusion_data = matrix_data['data']

        # Crear DataFrame de la matriz
        matrix_df = pd.DataFrame(matrix, index=labels, columns=labels)
        matrix_df.index.name = 'Gold_Label'
        matrix_df.columns.name = 'Predicted_Label'

        # Crear DataFrame de datos detallados
        details_df = pd.DataFrame(confusion_data)

        # Guardar en Excel con múltiples hojas
        output_path = Path(output_dir) / f'confusion_matrix_{prompt_name.lower().replace(" ", "_")}.xlsx'

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Hoja de la matriz
            matrix_df.to_excel(writer, sheet_name='Confusion_Matrix')

            # Hoja de detalles
            details_df.to_excel(writer, sheet_name='Detailed_Data', index=False)

            # Formatear la hoja de la matriz
            workbook = writer.book
            matrix_sheet = writer.sheets['Confusion_Matrix']

            # Aplicar formato a la matriz
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")

            # Formatear headers
            for cell in matrix_sheet[1]:
                cell.font = header_font
                cell.fill = header_fill

            for row in matrix_sheet.iter_rows(min_row=2, max_row=matrix_sheet.max_row,
                                              min_col=1, max_col=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill

        print(f"Matriz de confusión Excel guardada en: {output_path}")
        return output_path

    def analyze_annotations(self, gold_dir, system_dir, output_file=None, prompt_name="System"):
        tag_errors = defaultdict(lambda: {'fp': [], 'fn': [], 'tp': []})
        tag_sentences = defaultdict(int)
        all_confusion_data = []

        output_content = []
        output_content.append(f"MEDDOCAN Error Analysis Report - {prompt_name}")
        output_content.append("=" * 50)

        for filename in os.listdir(gold_dir):
            if filename.endswith(".xml"):
                gold_file = os.path.join(gold_dir, filename)
                system_file = os.path.join(system_dir, filename)

                if not os.path.exists(system_file):
                    continue

                # Parse annotations
                gold_ann = self.parse_i2b2_annotations(gold_file)
                sys_ann = self.parse_i2b2_annotations(system_file)

                # Count sentences
                num_sentences = gold_ann['text'].count('.') + gold_ann['text'].count('\n')
                for ann in gold_ann['annotations']:
                    tag_type = ann.get('subtype', ann['type'])
                    tag_sentences[tag_type] += num_sentences

                if output_file:
                    with open(output_file, 'a') as out_f:
                        self.compare_document_annotations(gold_ann, sys_ann, tag_errors, filename, out_f)
                else:
                    self.compare_document_annotations(gold_ann, sys_ann, tag_errors, filename)

                confusion_data = self.build_confusion_matrix_data(gold_ann, sys_ann, filename)
                all_confusion_data.extend(confusion_data)

        metrics = self.compute_metrics(tag_errors, tag_sentences)

        self.create_confusion_matrix(all_confusion_data, prompt_name)

        # Store results
        self.results[prompt_name] = {
            'metrics': metrics,
            'tag_errors': tag_errors,
            'tag_sentences': tag_sentences
        }

        return metrics

    def save_metrics_to_excel(self, metrics, output_path, prompt_name="System"):
        tags, precisions, recalls, f1s, leaks = [], [], [], [], []

        for tag, values in metrics.items():
            tags.append(tag)
            precisions.append(values['precision'])
            recalls.append(values['recall'])
            f1s.append(values['f1'])
            leaks.append(values['leak'])

        metrics_df = pd.DataFrame({
            'Tag': tags,
            'Precision': precisions,
            'Recall': recalls,
            'F1 Score': f1s,
            'Leak Score': leaks
        })

        excel_path = Path(output_path) / f"metrics_{prompt_name.lower().replace(' ', '_')}.xlsx"
        metrics_df.to_excel(excel_path, index=False)

        # Format Excel
        wb = load_workbook(excel_path)
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        alt_fill = PatternFill("solid", fgColor="DCE6F1")
        for row in range(2, ws.max_row + 1):
            if row % 2 == 0:
                for col in range(1, ws.max_column + 1):
                    ws[f"{get_column_letter(col)}{row}"].fill = alt_fill

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2

        wb.save(excel_path)
        return excel_path

    def compare_multiple_prompts(self, comparisons, output_dir="./results"):
        os.makedirs(output_dir, exist_ok=True)

        all_metrics = {}

        # Analizar cada prompt
        for prompt_name, (gold_dir, system_dir) in comparisons.items():
            print(f"Analizando {prompt_name}...")
            output_file = Path(output_dir) / f"error_report_{prompt_name.lower().replace(' ', '_')}.txt"
            if output_file.exists():
                output_file.unlink()

            metrics = self.analyze_annotations(gold_dir, system_dir, output_file=str(output_file),
                                               prompt_name=prompt_name)
            all_metrics[prompt_name] = metrics
            self.save_metrics_to_excel(metrics, output_dir, prompt_name)

            self.plot_confusion_matrix(prompt_name, output_dir)
            self.save_confusion_matrix_to_excel(prompt_name, output_dir)

        consolidated_data = []
        for prompt_name, metrics in all_metrics.items():
            for tag, values in metrics.items():
                consolidated_data.append({
                    'Prompt': prompt_name,
                    'Tag': tag,
                    'Precision': values['precision'],
                    'Recall': values['recall'],
                    'F1': values['f1'],
                    'Leak': values['leak'],
                    'TP': values['tp'],
                    'FP': values['fp'],
                    'FN': values['fn']
                })

        df_consolidated = pd.DataFrame(consolidated_data)
        df_consolidated.to_excel(Path(output_dir) / "consolidated_metrics.xlsx", index=False)

        self.create_comparison_plots(df_consolidated, output_dir)

        return df_consolidated

    def create_comparison_plots(self, df, output_dir):
        # Configurar estilo
        plt.style.use('default')
        sns.set_palette("husl")

        # Gráfico de F1 por tag y prompt
        plt.figure(figsize=(15, 8))
        pivot_f1 = df.pivot(index='Tag', columns='Prompt', values='F1')
        sns.heatmap(pivot_f1, annot=True, cmap='YlOrRd', fmt='.3f')
        plt.title('F1 Score Comparison by Tag and Prompt')
        plt.xticks(rotation=45)
        plt.yticks(rotation=0)
        plt.tight_layout()
        plt.savefig(Path(output_dir) / 'f1_heatmap.png', dpi=300, bbox_inches='tight')
        plt.show()

        # Gráfico de barras agrupadas para F1
        plt.figure(figsize=(15, 8))
        sns.barplot(data=df, x='Tag', y='F1', hue='Prompt')
        plt.title('F1 Score Comparison by Tag')
        plt.xticks(rotation=45)
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.tight_layout()
        plt.savefig(Path(output_dir) / 'f1_barplot.png', dpi=300, bbox_inches='tight')
        plt.show()

        # Gráfico de dispersión Precision vs Recall
        plt.figure(figsize=(12, 8))
        for prompt in df['Prompt'].unique():
            prompt_data = df[df['Prompt'] == prompt]
            plt.scatter(prompt_data['Precision'], prompt_data['Recall'],
                        label=prompt, alpha=0.7, s=100)

        plt.xlabel('Precision')
        plt.ylabel('Recall')
        plt.title('Precision vs Recall by Prompt')
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(Path(output_dir) / 'precision_recall_scatter.png', dpi=300, bbox_inches='tight')
        plt.show()

        # Resumen estadístico por prompt
        plt.figure(figsize=(12, 6))
        summary_stats = df.groupby('Prompt')[['Precision', 'Recall', 'F1']].mean()
        summary_stats.plot(kind='bar', ax=plt.gca())
        plt.title('Average Metrics by Prompt')
        plt.xticks(rotation=45)
        plt.legend()
        plt.tight_layout()
        plt.savefig(Path(output_dir) / 'average_metrics.png', dpi=300, bbox_inches='tight')
        plt.show()
        return summary_stats


# Funciones de conveniencia (funcionalidad original mantenida)
def analyze_single_prompt(gold_dir, system_dir, output_file=None, prompt_name="System"):
    """Función de conveniencia para analizar un solo prompt"""
    analyzer = MeddocanAnalyzer()
    return analyzer.analyze_annotations(gold_dir, system_dir, output_file, prompt_name)


def compare_prompts(comparisons, output_dir="./results"):
    """Función de conveniencia para comparar múltiples prompts"""
    analyzer = MeddocanAnalyzer()
    return analyzer.compare_multiple_prompts(comparisons, output_dir)